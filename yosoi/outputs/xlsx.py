"""XLSX output module (requires openpyxl)."""

import os
from pathlib import Path


def _cell_text(value: object) -> str:
    """Render one record value as cell text, with missing/None becoming a blank cell."""
    return '' if value is None else str(value)


def save_xlsx(filepath: str, url: str, domain: str, content: dict[str, object]) -> None:
    """Append one record to an XLSX workbook, creating it if it doesn't exist.

    Requires openpyxl. Install with: uv add openpyxl

    Args:
        filepath: Path to the .xlsx file
        url: Source URL
        domain: Domain name
        content: Extracted content dictionary

    Raises:
        ImportError: If openpyxl is not installed.

    """
    try:
        import openpyxl
    except (ImportError, TypeError):
        raise ImportError('openpyxl is required for XLSX output. Install it: uv add openpyxl') from None

    record = {'url': url, 'domain': domain, **content}

    if os.path.exists(filepath):
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        if ws is None:
            raise ValueError(f'{filepath} has no active worksheet')
        # Read existing header to align columns. Position matters: index i is column i+1,
        # so a blank header cell stays in the list and maps to a blank data cell.
        header: list[object] = [cell.value for cell in ws[1]]
        # Extend header with any new keys from the incoming record
        new_cols = [k for k in record if k not in header]
        if new_cols:
            for col_name in new_cols:
                header.append(col_name)
                ws.cell(row=1, column=len(header), value=col_name)
        row = [_cell_text(record.get(col) if isinstance(col, str) else None) for col in header]
        ws.append(row)
    else:
        Path(filepath).parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:  # pragma: no cover - a fresh Workbook always has an active sheet
            raise ValueError('new workbook has no active worksheet')
        ws.append(list(record.keys()))
        ws.append([_cell_text(v) for v in record.values()])

    wb.save(filepath)
