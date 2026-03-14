"""Tests for XLSX output module."""

import sys
from pathlib import Path

import pytest

URL = 'https://example.com/article'
DOMAIN = 'example.com'
CONTENT = {'headline': 'Test Article', 'author': 'Jane Doe'}


@pytest.fixture
def xlsx_save():
    pytest.importorskip('openpyxl', reason='openpyxl not installed')
    from yosoi.outputs.xlsx import save_xlsx

    return save_xlsx


def test_save_xlsx_creates_file(tmp_path, xlsx_save):
    filepath = str(tmp_path / 'results.xlsx')
    xlsx_save(filepath, URL, DOMAIN, CONTENT)
    assert Path(filepath).exists()


def test_save_xlsx_creates_workbook_with_header_row(tmp_path, xlsx_save):
    import openpyxl

    filepath = str(tmp_path / 'results.xlsx')
    xlsx_save(filepath, URL, DOMAIN, CONTENT)
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert 'url' in headers
    assert 'domain' in headers
    assert 'headline' in headers


def test_save_xlsx_writes_data_row(tmp_path, xlsx_save):
    import openpyxl

    filepath = str(tmp_path / 'results.xlsx')
    xlsx_save(filepath, URL, DOMAIN, CONTENT)
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    data_row = {headers[i]: ws.cell(2, i + 1).value for i in range(len(headers))}
    assert data_row['headline'] == 'Test Article'
    assert data_row['url'] == URL


def test_save_xlsx_appends_row_to_existing(tmp_path, xlsx_save):
    import openpyxl

    filepath = str(tmp_path / 'results.xlsx')
    xlsx_save(filepath, URL, DOMAIN, CONTENT)
    xlsx_save(filepath, 'https://example.com/page2', DOMAIN, {'headline': 'Second', 'author': 'Bob'})
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    # Row 1 = header, row 2 = first, row 3 = second
    assert ws.max_row == 3


def test_save_xlsx_no_duplicate_header_on_append(tmp_path, xlsx_save):
    import openpyxl

    filepath = str(tmp_path / 'results.xlsx')
    xlsx_save(filepath, URL, DOMAIN, CONTENT)
    xlsx_save(filepath, 'https://example.com/page2', DOMAIN, {'headline': 'Second', 'author': 'Bob'})
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    # Only one header row at row 1
    header_rows = [r for r in ws.iter_rows(values_only=True) if r[0] == 'url']
    assert len(header_rows) == 1


def test_save_xlsx_creates_parent_directory(tmp_path, xlsx_save):
    filepath = str(tmp_path / 'subdir' / 'nested' / 'results.xlsx')
    xlsx_save(filepath, URL, DOMAIN, CONTENT)
    assert Path(filepath).exists()


def test_save_xlsx_raises_import_error_without_openpyxl(monkeypatch, tmp_path):
    monkeypatch.setitem(sys.modules, 'openpyxl', None)
    import importlib

    import yosoi.outputs.xlsx as xlsx_mod

    importlib.reload(xlsx_mod)
    with pytest.raises(ImportError):
        xlsx_mod.save_xlsx(str(tmp_path / 'f.xlsx'), URL, DOMAIN, CONTENT)


def test_save_xlsx_none_values_stored_as_empty(tmp_path, xlsx_save):
    import openpyxl

    filepath = str(tmp_path / 'results.xlsx')
    xlsx_save(filepath, URL, DOMAIN, {'headline': None, 'author': 'Bob'})
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    data = {headers[i]: ws.cell(2, i + 1).value for i in range(len(headers))}
    # openpyxl reads empty cells as None; None values should not raise
    assert data['headline'] in (None, '')
    assert data['author'] == 'Bob'
